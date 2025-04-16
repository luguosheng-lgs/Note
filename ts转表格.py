#                                                                                          #
#              dynamic_Version需要增加，可以与软件版本同步                                    #
#              通常，需要等文档组回稿，更新翻译文件.ts后，再重新生成新的待翻译.xlsx              #


import os
import json
import sys
import re
from xml.dom import minidom
import xml.etree.ElementTree as ET
from xml.etree.ElementTree import Element
from collections import defaultdict
import argparse
import platform
from pathlib import Path
from os import path
from xml.dom.minidom import parse
import hashlib
import base64
import xlwt
import xlrd
import xlwings as xw
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl


Utils_Version = "V1.0.3"
dynamic_Version = "2.10.1"

py_file_dir = os.path.split(os.path.realpath(__file__))[0]
source_filepath = py_file_dir + "/../ConfigFiles/dynamic_config/"

tsFilePath = py_file_dir + "/ts/" 
xlsxFilePath = py_file_dir + "/xlsx/"

outFilePath = py_file_dir + "/TranslationOutFile/"
outTypeFilePath = py_file_dir + "/../ConfigFiles/dynamic_config/G4/"
referFilePath = py_file_dir + "/TranslationReferFile/"

file_name_list = ["robot_config_ui.json", "robot_state_ui.json"]

Gx_dirs = []
Ax_dirs = []
Vx_dirs = []
newObj = defaultdict(set)  #  '电源板序列号': {'A42-V1'}, '主控板序列号': {'A42-V1'},
newDict_en = dict()
newDict_ja = dict()
newDict_ko = dict()
newDict_es = dict()
newDict_robotFile = dict()
languageType = str()
msource = str()
mtranslation = str()
g_robotTypeList = list()


def extraction_tr(strdata, type):
    p1 = re.compile(r"tr\([\'](.*?)[\']", re.S)  # 最小匹配
    st_list = re.findall(p1, strdata)
    for st in st_list:
        newObj[st].add(type)


# obj:json文件内容, type:机型
def parseJson(obj, type):
    for k in obj:
        # 为list的时候
        if isinstance(obj, list):
            if isinstance(k, list):
                parseJson(k, type)
            elif isinstance(k, dict):
                parseJson(k, type)

        # 为dict的时候
        elif isinstance(obj, dict):
            if isinstance(obj[k], list):
                is_contain_tr = "_tr" in k
                if is_contain_tr == True:
                    for m in obj[k]:
                        newObj[m].add(type)
                elif k.startswith("on"):
                    for n in obj[k]:
                        extraction_tr(n, type)
                else:
                    parseJson(obj[k], type)
            elif isinstance(obj[k], dict):
                parseJson(obj[k], type)
            else:
                # 这一段 判断key
                is_contain_ch = "_tr" in k
                if is_contain_ch == True:
                    newObj[obj[k]].add(type)


def __indent(elem, level=0):
    i = "\n" + level * "\t"
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "\t"
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
        for elem in elem:
            __indent(elem, level + 1)
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = i


def return_key(val, tdict):
    for key in tdict.keys():
        if key == val:
            return key
    return "no"
def toXML(stData, parLanguage, parDict):
    TS = ET.Element("TS", {"version": "2.1", "language": parLanguage})
    tree = ET.ElementTree(TS)
    root = ET.Element("context")
    name = Element("name")
    name.text = "Null"
    root.append(name)
    for itme in stData:
        message = ET.Element("message")
        source = Element("source")
        source.text = itme
        stMode = list(set(stData[itme]))
        removePeelStr = ",".join(stMode)
        # gender = Element("location",{'filename': removePeelStr})
        # message.append(gender)
        message.append(source)
        if return_key(itme, parDict) == "no" or parDict[itme] == None:
            translation = Element("translation", {"type": "unfinished"})
            translation.text = ""
            message.append(translation)
        else:
            translation = Element("translation")
            translation.text = parDict[itme]
            message.append(translation)
        root.append(message)

    # 把旧的词条合并到新的文件中
    for referItem in parDict:
        if return_key(referItem, stData) == "no":
            message = ET.Element("message")
            source = Element("source")
            source.text = referItem
            message.append(source)
            translation = Element("translation")
            translation.text = parDict[referItem]
            message.append(translation)
            root.append(message)

    TS.append(root)
    __indent(TS)
    if os.path.exists(outFilePath) == False:
        os.mkdir(outFilePath)
    if os.path.exists(outFilePath) == False:
        return
    writePath = outFilePath + "HAI_Robot_Box_" + parLanguage + "_dynamic.ts"
    tree.write(writePath, encoding="utf-8", xml_declaration=True)
    fp = open(writePath, "r", encoding="utf-8")
    content = fp.readlines()
    fp.close()
    final_content = content[:1]  # 删除空键值
    final_content.append("<!DOCTYPE TS>\n")
    final_content.extend(content[1:])

    fp = open(writePath, "w", encoding="utf-8")
    fp.writelines(final_content)
    fp.close()
def readXML():
    for refer_file_name in os.listdir(referFilePath):
        if ".gitignore" in refer_file_name:
            continue
        refer_curr_path = referFilePath + refer_file_name
        if os.path.exists(refer_curr_path):
            treeRead = ET.parse(refer_curr_path)
            rootRead = treeRead.getroot()
            strRootAttrib = rootRead.attrib

            for strName in strRootAttrib:
                if strName == "language":
                    languageType = strRootAttrib[strName]
                    for child in rootRead:
                        for child1 in child:
                            msource = "NULL"
                            mtranslation = "NULL"
                            for child2 in child1:
                                if child2.tag == "source":
                                    msource = child2.text
                                if child2.tag == "translation":
                                    mtranslation = child2.text  # tag attrib

                                    if mtranslation == " " or mtranslation == None:
                                        for strName in child2.attrib:
                                            if strName == "type":
                                                tmpRobotType = child2.attrib[strName]
                                                if tmpRobotType == "unfinished":
                                                    mtranslation = "NULL"

                            if msource != "NULL" and mtranslation != "NULL":
                                if languageType == "en_US":
                                    newDict_en[msource] = mtranslation
                                elif languageType == "ko_KR":
                                    newDict_ko[msource] = mtranslation
                                elif languageType == "ja_JP":
                                    newDict_ja[msource] = mtranslation
                                elif languageType == "es_ES":
                                    newDict_es[msource] = mtranslation


def readWriteXML(languageEnum):
    for key in newObj.keys():
        if not newObj.get(key):
            del newObj[key]

    readXML()
    # if languageEnum & 1:
    #     toXML(newObj, "en_US", newDict_null)

    # if languageEnum & 2:
    #     toXML(newObj, "ko_KR", newDict_null)

    # if languageEnum & 4:
    #     toXML(newObj, "ja_JP", newDict_null)

    # if languageEnum & 8:
    #     toXML(newObj, "es_ES", newDict_null)


def check_system_platform():
    return platform.system()


class OrderedNamespace(argparse.Namespace):
    def __init__(self, **kwargs):
        self.__dict__["_order"] = [None]
        super().__init__(**kwargs)

    def __setattr__(self, attr, value):
        super().__setattr__(attr, value)
        if attr in self._order:
            self.__dict__["_order"].clear()
        self.__dict__["_order"].append(attr)

    def ordered(self):
        if self._order and self._order[0] is None:
            self._order.clear()
        return ((attr, getattr(self, attr)) for attr in self._order)
def languageSet(val):
    languageEnum = val
    readWriteXML(languageEnum)
    print("Execute successfully!")
    return True


def splitFiles():
    for file_name in os.listdir(tsFilePath):
        tmp_pathName = tsFilePath + file_name
        if os.path.exists(tmp_pathName):
            absName = Path(file_name).stem
            file_nameList = absName.split("_")
            file_nameList.reverse()
            if (
                analysisSplitFiles(
                    tmp_pathName, file_nameList[2] + "_" + file_nameList[1]
                )
                == False
            ):
                print("Exec failure !")
            else:
                print("Execute successfully!")


def analysisSplitFiles(tmp_pathName, parLanguage):
    print(parLanguage)
    global g_robotTypeList
    if os.path.exists(tmp_pathName):
        treeRead1 = ET.parse(tmp_pathName)
        rootRead1 = treeRead1.getroot()
        for child in rootRead1:
            for child1 in child:
                msource = "NULL"
                mtranslation = "NULL"
                for child2 in child1:
                    if child2.tag == "source":
                        msource = child2.text
                    if child2.tag == "translation":
                        mtranslation = child2.text  # tag attrib
                for child3 in child1:
                    if child3.tag == "location":
                        for strName in child3.attrib:
                            if strName == "filename":
                                robotType = child3.attrib[strName]
                                robotTypeList = robotType.split(",")
                                for robotName in robotTypeList:
                                    g_robotTypeList.append(robotName)
                        g_robotTypeList = list(set(g_robotTypeList))

        for tmpRobotName in g_robotTypeList:
            tmp_dict = {}
            for child in rootRead1:
                for child1 in child:
                    msource = "NULL"
                    mtranslation = "NULL"
                    for child2 in child1:
                        if child2.tag == "source":
                            msource = child2.text
                        if child2.tag == "translation":
                            mtranslation = child2.text  # tag attrib
                            if mtranslation == None:
                                for strName in child2.attrib:
                                    if strName == "type":
                                        tmpRobotType = child2.attrib[strName]
                                        if tmpRobotType == "unfinished":
                                            mtranslation = "NULL"
                    for child3 in child1:
                        if child3.tag == "location":
                            for strName in child3.attrib:
                                if strName == "filename":
                                    robotType = child3.attrib[strName]
                                    robotTypeList = robotType.split(",")
                                    if tmpRobotName in robotTypeList:
                                        if msource != "NULL":
                                            if parLanguage == "en_US":
                                                tmp_dict[msource] = mtranslation
                                            elif parLanguage == "ko_KR":
                                                tmp_dict[msource] = mtranslation
                                            elif parLanguage == "ja_JP":
                                                tmp_dict[msource] = mtranslation
                                            elif parLanguage == "es_ES":
                                                tmp_dict[msource] = mtranslation
                                            else:
                                                print(
                                                    '"'
                                                    + tmp_pathName
                                                    + '"'
                                                    + "file error!"
                                                )
                                                return False
            writeRobotTS(tmp_dict, parLanguage, tmpRobotName)
    return True
def writeRobotTS(dictData, parLanguage, robotType):
    TS = ET.Element("TS", {"version": "2.1", "language": parLanguage})
    tree = ET.ElementTree(TS)

    root = ET.Element("context")
    name = Element("name")
    root.append(name)
    for itme in dictData:
        message = ET.Element("message")

        source = Element("source")
        source.text = itme
        gender = Element("location", {"filename": robotType})
        message.append(gender)
        message.append(source)
        if dictData[itme] == "NULL":
            translation = Element("translation", {"type": "unfinished"})
            translation.text = None
            message.append(translation)
        else:
            translation = Element("translation")
            translation.text = dictData[itme]
            message.append(translation)
        root.append(message)
    TS.append(root)

    __indent(TS)
    robotTypeList = robotType.split("-")
    Path1 = outTypeFilePath + robotTypeList[0] + "/"
    Path2 = Path1 + robotTypeList[1] + "/"
    if os.path.exists(outTypeFilePath) == False:
        os.mkdir(outTypeFilePath)
    if os.path.exists(outTypeFilePath) == False:
        return

    if os.path.exists(Path1) == False:
        os.mkdir(Path1)
    if os.path.exists(Path1) == False:
        return

    if os.path.exists(Path2) == False:
        os.mkdir(Path2)
    if os.path.exists(Path2) == False:
        return

    # 保存到 ConfigFiles中各机型对应的目录
    writePath = Path2 + "HAI_Robot_Box_" + parLanguage + "_dynamic.ts"
    tree.write(writePath, encoding="utf-8", xml_declaration=True)
    fp = open(writePath, "r", encoding="utf-8")
    content = fp.readlines()
    fp.close()
    final_content = content[:1]  # 删除空键值
    final_content.append("<!DOCTYPE TS>\n")
    final_content.extend(content[1:])

    fp = open(writePath, "w", encoding="utf-8")
    fp.writelines(final_content)
    fp.close()


def checkTranslation():
    for file_name in os.listdir(tsFilePath):
        if os.path.exists(tsFilePath + file_name):
            index = 0
            print(tsFilePath + file_name)
            treeRead1 = ET.parse(tsFilePath + file_name)
            rootRead1 = treeRead1.getroot()
            for child in rootRead1:
                for child1 in child:
                    msource = "NULL"
                    mtranslation = "NULL"
                    for child2 in child1:
                        if child2.tag == "source":
                            msource = child2.text
                        if child2.tag == "translation":
                            mtranslation = child2.text  # tag attrib
                            for strName in child2.attrib:
                                if strName == "type":
                                    tmpRobotType = child2.attrib[strName]
                                    if (
                                        tmpRobotType == "unfinished"
                                        or mtranslation == None
                                    ):
                                        mtranslation = "NULL"
                                        print(msource + " = " + "unfinished")
                                        index += 1
            print('"' + file_name + '"' + " Total untranslated = {}".format(index))


# 加密算法 恺撒
def encrypt(lt):
    newlts = ""
    for index in lt:
        newlts = newlts + (chr(ord(index) + 2))
    return newlts


# 解密算法 恺撒
def decrypt(lt):
    newlts = ""
    for index in lt:
        newlts = newlts + (chr(ord(index) - 2))
    return newlts

# 加密算法 base64
def base64_encrypt(lt):
    return lt
    return base64.b64encode(lt.encode("utf-8")).decode("utf-8")


# 解密算法 base64
def base64_decrypt(lt):
    return lt
    return base64.b64decode(lt.encode("utf-8")).decode("utf-8")


# 加密文件，加密后保存在./EncryptionTranslationOutFile下，然后外发进行翻译
def encryptFile():
    if os.path.exists(xlsxFilePath) == False:
        os.mkdir(xlsxFilePath)
    if os.path.exists(xlsxFilePath) == False:
        return
    sourceFilePathList = list("")
    # 程序更新的ts文件
    sourceFilePathList.append(tsFilePath)
    for sourcefiles in sourceFilePathList:
        files = os.listdir(sourcefiles)
        for file in files:
            real_url = path.join(sourcefiles, file)
            print(real_url)
            if ".ts" in real_url:
                DOMTree = parse(real_url)
                collection = DOMTree.documentElement
                contexts = collection.getElementsByTagName("context")

                for context in contexts:
                    sourceData = (
                        context.getElementsByTagName("name")[0].childNodes[0].data
                    )
                    newData = base64_encrypt(sourceData)
                    context.getElementsByTagName("name")[0].childNodes[0].data = newData
                    messages = context.getElementsByTagName("message")
                    for message in messages:
                        locations = message.getElementsByTagName("location")
                        for location in locations:
                            filename = location.getAttribute("filename")
                            location.setAttribute("filename", base64_encrypt(filename))

                newfilePath = path.join(
                    xlsxFilePath, os.path.basename(real_url)
                )
                with open(newfilePath, "w", encoding="utf-8") as savePath:
                    DOMTree.writexml(savePath)
                    # 把加密后的xml 转 xlsx
                toxlsxByOpenpyxl(
                    newfilePath, os.path.splitext(newfilePath)[0] + ".xlsx"
                )
    return True

# xml 转 xlsx
def toxlsxByOpenpyxl(xmlPath, xlsxPath):
    print("xmlPath :" + xmlPath)
    print("xlsxPath :" + xlsxPath)

    DOMTree = parse(xmlPath)
    collection = DOMTree.documentElement
    contexts = collection.getElementsByTagName("context")

    # 新建工作薄
    wb = Workbook()

    index = 1
    sheet = wb.create_sheet("Seetong")
    del wb["Sheet"]
    for context in contexts:
        # 创建子sheet
        sourceData = context.getElementsByTagName("name")[0].childNodes[0].data

        sheet.cell(row=index, column=1).value = "[ContextName]"
        sheet.cell(row=index, column=2).value = sourceData
        sheet.cell(row=index, column=3).value = "[ContextName]行无需翻译"
        index = index + 1
        messages = context.getElementsByTagName("message")
        for message in messages:
            # 写入待翻译文本
            if message.getElementsByTagName("source")[0].hasChildNodes():
                source = message.getElementsByTagName("source")[0].childNodes[0].data
                sheet.cell(row=index, column=1).value = source
            else:
                sheet.cell(row=index, column=1).value = ""

            if message.getElementsByTagName("translation")[0].hasChildNodes():
                sheet.cell(row=index, column=2).value = (
                    message.getElementsByTagName("translation")[0].childNodes[0].data
                )
            else:
                if return_key(source, newDict_en) == "no" or newDict_en[source] == None:
                    sheet.cell(row=index, column=2).value = ""
                else:
                    sheet.cell(row=index, column=2).value = newDict_en[source]

            locationIndex = 1
            locations = message.getElementsByTagName("location")
            for location in locations:
                filename = location.getAttribute("filename")
                sheet.cell(row=index, column=2 + locationIndex).value = filename
                line = location.getAttribute("line")
                sheet.cell(row=index, column=3 + locationIndex).value = line
                locationIndex = locationIndex + 2
            index = index + 1

    wb.save(xlsxPath)
    wb.close()  # 关闭表格文件


# xml 转 xlsx
def toxlsxByXlwt(xmlPath, xlsxPath):
    print("xmlPath :" + xmlPath)
    print("xlsxPath :" + xlsxPath)

    DOMTree = parse(xmlPath)
    collection = DOMTree.documentElement
    contexts = collection.getElementsByTagName("context")

    # 新建工作薄
    new_workbook = xlwt.Workbook()

    for context in contexts:
        # 创建子sheet
        sourceData = context.getElementsByTagName("name")[0].childNodes[0].data
        sheet = new_workbook.add_sheet(sourceData)
        index = 0
        messages = context.getElementsByTagName("message")
        for message in messages:
            # 写入待翻译文本
            source = message.getElementsByTagName("source")[0].childNodes[0].data
            sheet.write(index, 0, source)

            if message.getElementsByTagName("translation")[0].hasChildNodes():
                sheet.write(
                    index,
                    1,
                    message.getElementsByTagName("translation")[0].childNodes[0].data,
                )
            else:
                sheet.write(index, 1, "")

            locationIndex = 1
            locations = message.getElementsByTagName("location")
            for location in locations:
                filename = location.getAttribute("filename")
                sheet.write(index, 1 + locationIndex, filename)
                line = location.getAttribute("line")
                sheet.write(index, 2 + locationIndex, line)
                locationIndex = locationIndex + 2
            index = index + 1
    new_workbook.save(xlsxPath)

# xlsx 转 xml
def xlsxToxml(xlsxPath, xmlPath):
    language = "en_US"
    if "_en" in xlsxPath:
        language = "en_US"
    if "_ko" in xlsxPath:
        language = "ko_KR"
    if "_ja" in xlsxPath:
        language = "ja_JP"
    if "_es" in xlsxPath:
        language = "es_ES"
    if "_ru" in xlsxPath:
        language = "ru_RU"

    newDoc = minidom.Document()
    TSNode = newDoc.createElement("TS")
    TSNode.setAttribute("version", "2.1")

    print("xlsxToxml :" + xlsxPath)
    bk = openpyxl.load_workbook(xlsxPath)
    sheet = bk.active
    minrow = sheet.min_row
    nrows = sheet.max_row
    mincol = sheet.min_column
    ncols = sheet.max_column

    hasChild = False
    contextNode = newDoc.createElement("context")
    for j in range(1, nrows + 1):
        AValue = sheet.cell(j, 1).value
        if AValue == "[ContextName]":

            if hasChild:
                TSNode.appendChild(contextNode)

            contextNode = newDoc.createElement("context")
            nameNode = newDoc.createElement("name")
            name = newDoc.createTextNode(sheet.cell(j, 2).value)
            nameNode.appendChild(name)
            contextNode.appendChild(nameNode)
            hasChild = True
            continue

        messageNode = newDoc.createElement("message")

        sourceNode = newDoc.createElement("source")
        sheetsource = sheet.cell(j, 1).value
        if isinstance(sheetsource, str):
            source = newDoc.createTextNode(sheet.cell(j, 1).value)
            sourceNode.appendChild(source)

        translationData = sheet.cell(j, 2).value
        if isinstance(translationData, str) and len(translationData) > 0:
            translationNode = newDoc.createElement("translation")
            translation = newDoc.createTextNode(translationData)
            translationNode.appendChild(translation)
        else:
            translationNode = newDoc.createElement("translation")
            translationNode.setAttribute("type", "unfinished")

        k = 1
        while k < (ncols - 2 + 1):
            filename = sheet.cell(j, k + 2).value
            line = sheet.cell(j, k + 3).value
            if filename != None:
                locationNode = newDoc.createElement("location")
                locationNode.setAttribute("filename", filename)
                locationNode.setAttribute("line", str(line))
                k = k + 2
                messageNode.appendChild(locationNode)
            else:
                break

        messageNode.appendChild(sourceNode)
        messageNode.appendChild(translationNode)
        contextNode.appendChild(messageNode)

    TSNode.appendChild(contextNode)
    newDoc.appendChild(TSNode)

    writePath = xmlPath
    with open(writePath, "w", encoding="utf-8") as savePath:
        print("save" + writePath)
        newDoc.writexml(savePath, "", " " * 4, "\n", "UTF-8")

    # TS = ET.Element(newDoc.toprettyxml())
    # __indent(TS)
    # tree = ET.ElementTree(TS)
    # writePath = xmlPath
    # tree.write(writePath, encoding='utf-8', xml_declaration=False)

    fp = open(writePath, "r", encoding="utf-8")
    content = fp.readlines()
    fp.close()
    final_content = content[:1]  # 删除空键值
    final_content.append("<!DOCTYPE TS>\n")
    final_content.extend(content[1:])

    fp = open(writePath, "w", encoding="utf-8")
    fp.writelines(final_content)
    fp.close()

# 解密文件，外部翻译完成后，进行解密，解密到对应文件路径
def decryptFile():
    if os.path.exists(tsFilePath) == False:
        os.mkdir(tsFilePath)
    if os.path.exists(tsFilePath) == False:
        return
    files = os.listdir(xlsxFilePath)

    # 把 xlsx 转回 xml文件
    file = "seetong_ru.xlsx"
    real_url = path.join(xlsxFilePath, file)
    print(real_url)
    if ".xlsx" in real_url:
        xlsxToxml(real_url, os.path.splitext(real_url)[0] + ".ts")


def init_opt_parse():
    _parse = argparse.ArgumentParser(description="Internationalized data extraction")
    _parse.add_argument(
        "-c",
        "--check",
        action="store_true",
        dest="check",
        help="check translation type= unfinished",
    )
    _parse.add_argument(
        "-e",
        action="store",
        type=int,
        dest="extraction_data",
        help="Types of translation languages en_US = 1,ko_KR = 2,ja_JP = 4,es_ES = 8; \n\t \
                              example:en_US | ko_KR -> xxxx.py  -e 3 ;\
                                      en_US | ko_KR |ja_JP -> xxxx.py  -e 7 ;",
    )
    _parse.add_argument(
        "-s",
        "--split",
        action="store_true",
        dest="Split",
        help="Split translation files",
    )
    _parse.add_argument(
        "-enc",
        "--encrypt",
        action="store_true",
        dest="encrypt",
        help="encrypt translation files",
    )
    _parse.add_argument(
        "-dec",
        "--decrypt",
        action="store_true",
        dest="decrypt",
        help="decrypt translation files",
    )
    _parse.add_argument(
        "-v",
        "--version",
        action="version",
        version="%(prog)s {0}".format(Utils_Version),
    )
    return _parse

# 常规步骤：
# 1.执行：TranslationExtractionTool.py  -e 7
# 2.把EncryptionTranslationOutFile文件夹中的文件外发翻译
# 3.翻译好的文件，放在TranslationFinished文件夹
# 4.执行：TranslationExtractionTool.py  -s
if __name__ == "__main__":
    _parse = init_opt_parse()
    if len(sys.argv[1:]) < 1:
        print("[Error]: invalid argument for command line")
        _parse.print_help()
        sys.exit(1)
    args = _parse.parse_args(sys.argv[1:], namespace=OrderedNamespace())
    system_type = check_system_platform()

    for arg, val in args.ordered():
        if arg == "check":
            if not checkTranslation():
                sys.exit(1)
        if (
            arg == "extraction_data"
        ):  # 提取+加密，最后输出到EncryptionTranslationOutFile文件夹，然后把ts文件进行外发翻译
            print(val)
            # if languageSet(val):
            if not encryptFile():
                sys.exit(1)
        if (
            arg == "Split"
        ):  # 解密+拆分，把TranslationFinished文件夹中翻译好的ts文件解密并拆分到各自目标路径
            decryptFile()
            # if not splitFiles():
            #     sys.exit(1)
        if arg == "encrypt":
            sys.exit(1)
        if arg == "decrypt":
            sys.exit(1)

        elif arg == "version":
            print("version:" + Utils_Version)